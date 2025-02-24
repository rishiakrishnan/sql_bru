import os
import re
import zipfile
import xml.etree.ElementTree as ET
from docx import Document
from pptx import Presentation
from openpyxl import load_workbook

def extract_metadata(file_path):
    """Extract metadata from .docx, .pptx, .xlsx files"""
    metadata = {}

    try:
        if file_path.endswith(".docx"):
            doc = Document(file_path)
            metadata["Title"] = doc.core_properties.title
            metadata["Author"] = doc.core_properties.author
            metadata["Comments"] = doc.core_properties.comments

        elif file_path.endswith(".pptx"):
            ppt = Presentation(file_path)
            metadata["Title"] = ppt.core_properties.title
            metadata["Author"] = ppt.core_properties.author
            metadata["Comments"] = ppt.core_properties.comments

        elif file_path.endswith(".xlsx"):
            wb = load_workbook(file_path)
            metadata["Sheet Names"] = wb.sheetnames
            metadata["Properties"] = wb.properties.__dict__

    except Exception as e:
        metadata["Error"] = f"Failed to extract metadata: {e}"

    return metadata


def scan_xml_yaml(file_path):
    """Scan XML/YAML files for hidden scripts or suspicious content"""
    suspicious_patterns = [
        r"<script>.*?</script>",  # Embedded scripts
        r"<!ENTITY\s+.*?SYSTEM",  # XXE (External Entity Injection)
        r"eval\(",                # Code execution
        r"document\.cookie",       # JavaScript Cookie Stealing
        r"base64_decode\(",        # Obfuscated Code
        r"python:",                # YAML Python Execution
        r"!!python",               # YAML Deserialization Attack
    ]

    try:
        if file_path.endswith((".xml", ".yaml", ".yml")):
            with open(file_path, "r", errors="ignore") as file:
                content = file.read()
                for pattern in suspicious_patterns:
                    if re.search(pattern, content, re.IGNORECASE):
                        return True
    except Exception as e:
        print(f"Error scanning {file_path}: {e}")

    return False
